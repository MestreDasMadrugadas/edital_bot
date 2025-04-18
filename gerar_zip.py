import os
from zipfile import ZipFile

files = {
    "app.py": '''
from flask import Flask, request
from utils import gerar_planilha_personalizada, enviar_email_com_anexo

app = Flask(__name__)

@app.route('/webhook', methods=['POST'])
def webhook():
    data = request.json
    nome = data['purchase']['buyer']['name']
    cpf = data['purchase']['buyer']['document'].replace('.', '').replace('-', '')
    email = data['purchase']['buyer']['email']

    planilha = gerar_planilha_personalizada(nome, cpf)
    enviar_email_com_anexo(nome, email, planilha)
    return "Planilha enviada com sucesso!", 200

if __name__ == '__main__':
    app.run(debug=True)
''',

    "utils.py": '''
from openpyxl import load_workbook
import os
import smtplib
from email.message import EmailMessage

EMAIL_REMETENTE = "mestredasmadrugadas@gmail.com"
SENHA_APP = "SUA_SENHA_DE_APP_AQUI"

def gerar_planilha_personalizada(nome, cpf):
    caminho_modelo = 'edital_modelo.xlsx'
    wb = load_workbook(caminho_modelo)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.replace('{{nome}}', nome).replace('{{cpf}}', cpf)
    nome_arquivo = f'edital_{nome.replace(" ", "_")}.xlsx'
    os.makedirs('editais', exist_ok=True)
    caminho_final = os.path.join('editais', nome_arquivo)
    wb.save(caminho_final)
    return caminho_final

def enviar_email_com_anexo(nome, destinatario, arquivo_xlsx):
    msg = EmailMessage()
    msg['Subject'] = '📘 Seu edital personalizado está pronto!'
    msg['From'] = f"Mestre das Madrugadas <{EMAIL_REMETENTE}>"
    msg['To'] = destinatario
    corpo_email = f"""
    Olá {nome},

    Obrigado por sua compra! 🎉

    Segue em anexo sua planilha personalizada com o edital verticalizado.

    Bons estudos e até a aprovação! 💪
    — Equipe Mestre das Madrugadas
    """
    msg.set_content(corpo_email)
    with open(arquivo_xlsx, 'rb') as f:
        msg.add_attachment(
            f.read(),
            maintype='application',
            subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=os.path.basename(arquivo_xlsx)
        )
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_REMETENTE, SENHA_APP)
        smtp.send_message(msg)
''',

    "requirements.txt": "Flask\nopenpyxl",
    "README.txt": "Instruções:\n1. Substitua a senha de app em utils.py\n2. Rode localmente com Flask\n3. Teste o endpoint /webhook\n4. Suba no Render com gunicorn app:app"
}

os.makedirs("editais", exist_ok=True)

with ZipFile("edital_bot_render_smtp.zip", "w") as zipf:
    for filename, content in files.items():
        with open(filename, "w", encoding="utf-8") as f:
            f.write(content.strip())
        zipf.write(filename)
    zipf.write("requirements.txt")
    zipf.write("README.txt")
    zipf.write("editais")