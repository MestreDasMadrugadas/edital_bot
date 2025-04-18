from openpyxl import load_workbook
import os
import smtplib
from email.message import EmailMessage

EMAIL_REMETENTE = "mestredasmadrugadas@gmail.com"
SENHA_APP = "SUA_SENHA_DE_APP_AQUI"

def gerar_planilha_personalizada(nome, cpf):
    caminho_modelo = 'edital_modelo.xlsx'
    wb = load_workbook(caminho_modelo)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.replace('{{nome}}', nome).replace('{{cpf}}', cpf)
    nome_arquivo = f'edital_{nome.replace(" ", "_")}.xlsx'
    os.makedirs('editais', exist_ok=True)
    caminho_final = os.path.join('editais', nome_arquivo)
    wb.save(caminho_final)
    return caminho_final

def enviar_email_com_anexo(nome, destinatario, arquivo_xlsx):
    msg = EmailMessage()
    msg['Subject'] = '📘 Seu edital personalizado está pronto!'
    msg['From'] = f"Mestre das Madrugadas <{EMAIL_REMETENTE}>"
    msg['To'] = destinatario
    corpo_email = f"""
    Olá {nome},

    Obrigado por sua compra! 🎉

    Segue em anexo sua planilha personalizada com o edital verticalizado.

    Bons estudos e até a aprovação! 💪
    — Equipe Mestre das Madrugadas
    """
    msg.set_content(corpo_email)
    with open(arquivo_xlsx, 'rb') as f:
        msg.add_attachment(
            f.read(),
            maintype='application',
            subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=os.path.basename(arquivo_xlsx)
        )
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_REMETENTE, SENHA_APP)
        smtp.send_message(msg)